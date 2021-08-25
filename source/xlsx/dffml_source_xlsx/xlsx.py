from dffml.record import Record
from dffml.source.source import *
from dffml.util.entrypoint import entrypoint
from dffml.base import config
from dffml import (
    config,
    entrypoint,
    Record,
    FileSource,
    MemorySource,
)
from openpyxl import Workbook, load_workbook


@config
class XLSXSourceConfig:
    filename: str
    readwrite: bool = False
    allowempty: bool = False
    sheet: str = None
    FIRST_ROW_CONTAINS_COLUMN_NAMES: bool = True


@entrypoint("xlsx")
class XLSXSource(FileSource, MemorySource):
    """
    Source for interacting with Excel files (.xlsx, .xlsm, .xltx, .xltm)
    """

    CONFIG = XLSXSourceConfig

    async def load_fd(self, fileobj):
        self.wb = load_workbook(fileobj.name)  # Load the excel workbook

        # Check if a sheet is specified, if not select the active sheet.
        if not self.config.sheet:
            ws = self.wb.active
        else:
            ws = self.wb[self.config.sheet]

        # Record column names for convinience.
        if self.config.FIRST_ROW_CONTAINS_COLUMN_NAMES:
            self.columns = {
                cell.value: cell.column_letter
                for cell in ws[1][: ws.max_column + 1]
                if cell.value
            }
        else:
            self.columns = {
                cell.column_letter: cell.column_letter
                for cell in ws[1][: ws.max_column + 1]
                if cell.value
            }

        # Load data into memory by iterating over rows in the sheet.
        self.mem = {
            str(row): Record(
                key=str(row),
                data={
                    "features": {
                        feature: ws[self.columns[feature] + str(row + 1)].value
                        for feature in self.columns
                    }
                },
            )
            for row in range(1, ws.max_row)
        }
        self.logger.debug(
            "Loaded %s records from XLSX file: %s", len(self.mem), fileobj.name
        )

    async def dump_fd(self, fileobj):
        # Initialise self.columns to an empty dictionary for new files.
        try:
            self.columns
        except AttributeError:
            self.columns = {}

        # If file does not exist, create a blank workbook.
        try:
            self.wb
        except (NameError, AttributeError):
            try:
                self.wb = load_workbook(fileobj.name)
            except:
                self.wb = Workbook()

        # Create blank sheet if no sheets available in the file.
        if len(self.wb.sheetnames) == 0:
            self.wb.create_sheet(self.config.sheet)

        # Set desired sheet as active.
        if not self.config.sheet:
            ws = self.wb.active
        else:
            try:
                ws = self.wb[self.config.sheet]
            except KeyError:
                self.wb.create_sheet(self.config.sheet)
                ws = self.wb[self.config.sheet]

        # Iterate over self.mem to update rows into the file.
        for record_key in self.mem.keys():
            data = self.mem[record_key].features()
            # Iterate over each column in the current record.
            for column in data.keys():
                # If new column is added in one of the records, update the same in the file and self.columns
                if column not in self.columns.keys():
                    cur_col = ws.max_column
                    if ws.cell(row=1, column=cur_col).value != None:
                        cur_col += 1
                    ws.cell(row=1, column=cur_col).value = column
                    self.columns[column] = ws.cell(
                        row=1, column=cur_col
                    ).column_letter
                # Set value of each cell in the row according to record feature values.
                ws[
                    self.columns[column]
                    + str(int(self.mem[record_key].key) + 1)
                ] = data[column]
        self.wb.save(filename=fileobj.name)
        self.logger.debug(f"{self.config.filename} written")
        self.logger.debug("%r saved %d records", self, len(self.mem))
